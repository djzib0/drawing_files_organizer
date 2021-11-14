import os
import sys
import shutil
from openpyxl.utils import column_index_from_string


from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QPushButton, QComboBox,
                             QCheckBox, QFrame, QVBoxLayout, QGridLayout, QLineEdit,
                             QMessageBox
                             )
from PyQt5.QtCore import pyqtSlot, QRect

def set_xlsx_files_list():
    """Creates list of xlsx files."""
    files = os.listdir()
    xlsx_files = []
    for file in files:
        if "xlsx" in file:
            xlsx_files.append(file)
    return xlsx_files

def check_file_extension_len(file_name):
    """Checks how long is extension string"""
    list_ = file_name.split('.')
    return len(list_[1])

def create_key(list_):
    """Creates string key with coordinates for copying files."""
    key = str(list_)
    return str(key)

def create_short_key(list_):
    """Creates string key with coordinates from [0:-1]"""
    key = str(list_[0:-1])
    return str(key)

def test_print():
    print("Ni chuja ale to działa.")

def create_folders_and_copy_files(tree, part, name, start):
    import openpyxl
    # open excel file
    book = openpyxl.load_workbook('input.xlsx')
    sheet = book.active
    last_row = int(sheet.max_row) # sets last row with data

    drawings = [] # list keeps names of files to be copied
    EXTENSIONS = ['pdf', 'dwg'] # list of file extensions (TO BE MODIFIED)
    files_to_copy = []
    file_list = os.listdir() # creates file list in folder

    checked_row = [0]
    below_row = [0]
    current_path = os.getcwd()
    paths = {'[0]': current_path + '\\new_folder'}
    paths_short = {'[0]': current_path + '\\new_folder'}

    last_tree_col = tree # value of last column with coorindates
    part_col = part  # ta wartość będzie domyślna, ale z możliwością modyfikacji, podaje numer kolumny z part number
    name_col = name # ta wartość będzie domyślna, ale z możliwością modyfikacji, podaje numer kolumny z nazwą części
    row_counter = start # starting row
    col_counter = 1 # starting column
    key = ''

    # adding part/drawing numbers to the list

    # creating list of files valid to be copied

    for file in file_list:
            files_to_copy.append(file)

    print(files_to_copy)

    # analizing excel file - creating paths for folders and subfolders
    while row_counter < last_row:
        checked_row = [0]
        below_row = [0]
        for col in range(1, last_tree_col):
            if sheet.cell(row=row_counter, column=col).value != None: # sprawdza czy wartość komórki jest None
                checked_row.append(sheet.cell(row=row_counter, column=col).value) # jeżeli nie to dodaje do listy
            if sheet.cell(row=row_counter+1, column=col).value != None: # sprawdza czy wartość komórki jest None
                below_row.append(sheet.cell(row=row_counter+1, column=col).value) # jeżeli nie to dodaje do listy

        print(f'test checked {checked_row}')
        print(f'test_below {below_row}')
        if len(checked_row) < len(below_row): # sprawdza czy lista podrzędna jest dłuższa
            #checked_row.append('0') # dodaje do listy pierwszy element, który będzie koordynatą dla bieżącego katalogu
            print('dodaję koordynaty do słownika i tworzę katalog') # jeżeli tak to dodaje koordynaty do słownika i tworzy katalog
            # sprawdza czy klucz już występuje w słowniku, jeżeli tak to nie dodaje nowego.
            print(f'checked row to {checked_row}')
            key = create_key(checked_row)
            print('dodaję key')
            print(f'key to {key}')
            print(f'to jest checked_row przed short key {checked_row}')
            short_key = create_short_key(checked_row)
            print(f'to jest short key  {short_key}')
            part_number = str(sheet.cell(row=row_counter, column=part_col).value)
            part_name = str(sheet.cell(row=row_counter, column=name_col).value)

            print(f'checked_row to: {checked_row}')

            print(f'\nTo jest key {key}')
            print(f'To jest short key {short_key}\n')

            if key not in paths.keys():
                paths[key] = str(paths[short_key]) + '\\' + part_number + ' ' + part_name

        row_counter += 1
        #checked_row = []
        below_row = []

    # creating folders
    for i, path in paths.items():
        print(f'to jest klucz {i}')
        print(f'to jest ścieżka {path}')
        try:
            if i != '[0]':
                os.makedirs(str(path))
            else:
                print('to jest katalog startowy')
        except FileExistsError:
            print('Ten folder już istnieje')

    # copying files
    row_counter = 2
    checked_row = [0]

    print(paths)
    print(paths_short)
    while row_counter < last_row:
        checked_row = [0]
        below_row = [0]
        for col in range(1, last_tree_col):
            if sheet.cell(row=row_counter, column=col).value != None: # sprawdza czy wartość komórki jest None
                checked_row.append(sheet.cell(row=row_counter, column=col).value) # jeżeli nie to dodaje do listy
            if sheet.cell(row=row_counter+1, column=col).value != None: # sprawdza czy wartość komórki jest None
                below_row.append(sheet.cell(row=row_counter+1, column=col).value) # jeżeli nie to dodaje do listy
        drawing = sheet.cell(row_counter, part_col).value
        print(f'to jest {drawing}')
        for file in files_to_copy:
            if drawing in file:
                if len(checked_row) < len(below_row):
                    to_be_copied = str(current_path) + '\\' + str(file)
                    destination_folder = paths[create_short_key(below_row)]
                    print(destination_folder)
                    shutil.copy(to_be_copied, destination_folder)
                else:
                    print("tak jest w liście")
                    to_be_copied = str(current_path) + '\\' + str(file)
                    destination_folder = paths[create_short_key(checked_row)]
                    print(f'to jest to be_copied {to_be_copied}')
                    print(f'to jest kopiowany file {drawing}')
                    print(str(current_path) + '\\' + str(file))
                    shutil.copy(to_be_copied, destination_folder)

        row_counter += 1

def create_short_name(text, letters):
    """Creates shorter name for a given string"""
    item_list = text.split(" ")
    new_text = ""
    print(item_list)
    for item in item_list:
        if len(item) <= letters:
            new_text += item + " "
        else:
            new_text += item[0:letters] + " "


xlsx_files = set_xlsx_files_list()

class MainPage(QWidget):
    def __init__(self, title=" "):

        super().__init__() # inherit init of QWidget
        self.title = title
        self.left = 250
        self.top = 250
        self.width = 200
        self.height = 150

        self.widget()

    def widget(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        # setting main app layout
        main_layout = QVBoxLayout()
        settings_grid_layout = QGridLayout()
        self.setLayout(main_layout)
        self.setLayout(settings_grid_layout)

        self.choose_file_lbl = QLabel("Wybierz plik", self)
        main_layout.addWidget(self.choose_file_lbl)  # adds label to grid layout
        self.choose_file_combobox = QComboBox(self)
        for file in xlsx_files:
            self.choose_file_combobox.addItem(file)  # adds files to combobox list.
        main_layout.addWidget(self.choose_file_combobox)  # adds combobox to grid layout

        main_layout.addLayout(settings_grid_layout) # adds grid layout to main Vertical layout.

        self.set_last_tree_column_lbl = QLabel("Podaj ostatnią kolumnę drzewka", self)
        settings_grid_layout.addWidget(self.set_last_tree_column_lbl, 1, 0)  # adds label to grid layout
        self.set_last_tree_column_entry = QLineEdit(self)
        settings_grid_layout.addWidget(self.set_last_tree_column_entry, 1, 1) # adds LineEdit to grid layout

        self.set_part_column_lbl = QLabel("Podaj kolumnę z numerem części", self)
        settings_grid_layout.addWidget(self.set_part_column_lbl, 2, 0)  # adds label to grid layout
        self.set_part_column_entry = QLineEdit(self)
        settings_grid_layout.addWidget(self.set_part_column_entry, 2, 1)  # adds LineEdit to grid layout

        self.set_name_column_lbl = QLabel("Podaj kolumnę z nazwą części", self)
        settings_grid_layout.addWidget(self.set_name_column_lbl, 3, 0)  # adds label to grid layout
        self.set_name_column_entry = QLineEdit(self)
        settings_grid_layout.addWidget(self.set_name_column_entry, 3, 1)  # adds LineEdit to grid layout

        self.set_start_row_lbl = QLabel("Podaj wiersz, od którego należy zacząć", self)
        settings_grid_layout.addWidget(self.set_start_row_lbl, 4, 0)  # adds label to grid layout
        self.set_start_row_entry = QLineEdit(self)
        settings_grid_layout.addWidget(self.set_start_row_entry, 4, 1)  # adds LineEdit to grid layout

        self.ask_to_short_names_lbl = QLabel("Skrócić nazwy plików?", self)
        settings_grid_layout.addWidget(self.ask_to_short_names_lbl, 5, 0)  # adds label to grid layout
        self.ask_to_short_names_check = QCheckBox("Tak", self)
        settings_grid_layout.addWidget(self.ask_to_short_names_check, 5, 1)  # adds checkbox to grid layout
        self.ask_to_short_names_check.stateChanged.connect(self.ask_to_short_names)

        self.ask_how_many_letters_lbl = QLabel("Ile liter na słowo?", self)
        settings_grid_layout.addWidget(self.ask_how_many_letters_lbl, 6, 0)  # adds label to grid layout
        self.ask_how_many_letters_entry = QLineEdit(self)
        self.ask_how_many_letters_entry.setEnabled(False)
        settings_grid_layout.addWidget(self.ask_how_many_letters_entry, 6, 1)  # adds LineEdit to grid layout

        self.ask_to_create_unused_parts_lbl = QLabel("Stworzyć listę braków?", self)
        settings_grid_layout.addWidget(self.ask_to_create_unused_parts_lbl, 7, 0)  # adds label to grid layout
        self.ask_to_create_unused_check = QCheckBox("Tak", self)
        settings_grid_layout.addWidget(self.ask_to_create_unused_check, 7, 1)  # adds checkbox to grid layout
        self.ask_to_short_names_check.stateChanged.connect(self.ask_to_short_names)

        self.run_script_bttn = QPushButton("Uruchom", self)
        self.run_script_bttn.setFixedWidth(100)
        self.run_script_bttn.setFixedHeight(30)
        self.run_script_bttn.clicked.connect(self.run_code)
        settings_grid_layout.addWidget(self.run_script_bttn, 8, 0) # add button to grid layout

        self.show()

    def ask_to_short_names(self):
        if self.ask_to_short_names_check.isChecked():
            self.ask_how_many_letters_entry.setEnabled(True)
        else:
            self.ask_how_many_letters_entry.setEnabled(False)
           # self.show()

    @pyqtSlot()
    def check_errors(self):
        """Starts script - sorts and copies files."""
        self.errors_counter = 0
        self.last_tree_col_coordinates = '' # set variable to show it's string
        self.last_tree_column = 0 # set variable to show it's integer
        self.part_col_coordinates = '' # set variable to show it's string
        self.part_column = 0 #set variable to show it's integer
        self.name_col_coordinates = ''  # set variable to show it's string
        self.name_column = 0  # set variable to show it's integer
        self.start_row_coordinates = 0 # set variable to show it's integer
        self.start_row = 0 # set variable to show it's an integer

        # Checks if value is a letter, if not it gives an error.
        if self.set_last_tree_column_entry.text().lower().isalpha():
            self.last_tree_col_coordinates = self.set_last_tree_column_entry.text()
            self.last_tree_column = column_index_from_string(self.last_tree_col_coordinates) + 1  # add 1 to set range from (0, x)
            print(f'To jest indeks last_tree + 1: {self.last_tree_column}')
        else:
            self.errors_counter += 1
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Błąd wartości 'Kolumny Drzewka'")
            msg.setInformativeText('Nazwa kolumny musi zawierać same litery.')
            msg.setWindowTitle("Error")
            msg.exec_()

        # Checks if value is a letter, if not it gives an error.
        if self.set_part_column_entry.text().lower().isalpha():
            self.part_col_coordinates = self.set_part_column_entry.text()
            self.part_column = column_index_from_string(self.part_col_coordinates)
            print(f'To jest part column: {self.part_column}')
        else:
            self.errors_counter += 1
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Błąd wartości 'Kolumny z numerem części'")
            msg.setInformativeText('Nazwa kolumny musi zawierać same litery.')
            msg.setWindowTitle("Error")
            msg.exec_()

        # Checks if value is a letter, if not it gives an error.
        if self.set_name_column_entry.text().lower().isalpha():
            self.name_col_coordinates = self.set_name_column_entry.text()
            self.name_column = column_index_from_string(self.name_col_coordinates)
            print(f'To jest name column: {self.name_column}')
        else:
            self.errors_counter += 1
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Błąd wartości 'Kolumny z nazwą części'")
            msg.setInformativeText('Nazwa kolumny musi zawierać same litery.')
            msg.setWindowTitle("Error")
            msg.exec_()

        # # Checks if value is a letter, if not it gives an error.
        # if self.set_part_column_entry.text().lower().isalpha():
        #     self.part_col_coordinates = self.set_part_column_entry.text()
        #     self.part_column = column_index_from_string(self.part_col_coordinates)
        #     print(self.part_column)
        # else:
        #     msg = QMessageBox()
        #     msg.setIcon(QMessageBox.Critical)
        #     msg.setText(f"Błąd wartości 'Kolumny z numerem części'")
        #     msg.setInformativeText('Nazwa kolumny musi zawierać same litery.')
        #     msg.setWindowTitle("Error")
        #     msg.exec_()

        # Checks if value is a number, if not it gives an error.
        if self.set_start_row_entry.text().isnumeric():
            self.start_row_coordinates = self.set_start_row_entry.text()
            self.start_row = int(self.start_row_coordinates)
            print(f'to jest wiersz początkowy: {self.start_row}')
        else:
            self.errors_counter += 1
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Błąd wartości 'Wiersza Początkowego'")
            msg.setInformativeText('Nazwa wiersza musi zawierać same cyfry.')
            msg.setWindowTitle("Error")
            msg.exec_()

        print(f"To jest licznik błędów: {self.errors_counter}")
        print(f"To jest self combobox: {self.choose_file_combobox.currentText()}")

        return self.errors_counter




    def check_col_coord_is_valid(self, coordinate):
        """Checks if column coordinate is valid"""
        return coordinate.isalpha()

    @pyqtSlot()
    def run_code(self):
        try:
            if self.check_errors() == 0:
                create_folders_and_copy_files(int(self.last_tree_column),
                                              int(self.part_column),
                                              int(self.name_column),
                                              int(self.start_row),

                                              )
                print(int(self.last_tree_column))
                print(int(self.part_column))
                print(int(self.name_column))
                print(int(self.start_row))
        except:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Oops")
            msg.setInformativeText('Gibony chyba coś popsuły.\n Sprawdź dane lub plik wejściowy.')
            msg.setWindowTitle("Error")
            msg.exec_()


def main():
    app = QApplication(sys.argv)
    w = MainPage(title="Aplikacja Piotera")
    sys.exit(app.exec_())

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
