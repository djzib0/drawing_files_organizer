# App for creating folders and copying files according given list

import os
import shutil
import openpyxl


def check_file_extension_len(file_name):
    """Checks how long is extension string"""
    list_ = file_name.split('.')
    return len(list_[1])

def create_key(list_):
    """Creates string key with coordinates for copying files."""
    key = str(list_)
    return key

def create_short_key(list_):
    key = str(list_[0:-1])
    return key

# open excel file
book = openpyxl.load_workbook('input2.xlsx')
sheet = book.active
last_row = int(sheet.max_row) # sets last row with data

drawings = [] # list keeps names of files to be copied
EXTENSIONS = ['pdf', 'dwg'] # list of file extensions (TO BE MODIFIED)
files_to_copy = []
file_list = os.listdir() # creates file list in folder

checked_row = [0]
below_row = [0]
current_path = os.getcwd()
paths = {'[0]': current_path + '\\new_folder'}
paths_short = {'[0]': current_path + '\\new_folder'}

part_col = 7  # ta wartość będzie domyślna, ale z możliwością modyfikacji, podaje numer kolumny z part number
name_col = 8 # ta wartość będzie domyślna, ale z możliwością modyfikacji, podaje numer kolumny z nazwą części
row_counter = 2 # starting row
col_counter = 1 # starting column
key = ''

# adding part/drawing numbers to the list

# creating list of files valid to be copied

for file in file_list:
        files_to_copy.append(file)

print(files_to_copy)

# analizing excel file - creating paths for folders and subfolders
while row_counter < last_row:
    checked_row = [0]
    below_row = [0]
    for col in range(1, 6):
        if sheet.cell(row=row_counter, column=col).value != None: # sprawdza czy wartość komórki jest None
            checked_row.append(sheet.cell(row=row_counter, column=col).value) # jeżeli nie to dodaje do listy
        if sheet.cell(row=row_counter+1, column=col).value != None: # sprawdza czy wartość komórki jest None
            below_row.append(sheet.cell(row=row_counter+1, column=col).value) # jeżeli nie to dodaje do listy

    print(f'test checked {checked_row}')
    print(f'test_below {below_row}')
    if len(checked_row) < len(below_row): # sprawdza czy lista podrzędna jest dłuższa
        #checked_row.append('0') # dodaje do listy pierwszy element, który będzie koordynatą dla bieżącego katalogu
        print('dodaję koordynaty do słownika i tworzę katalog') # jeżeli tak to dodaje koordynaty do słownika i tworzy katalog
        # sprawdza czy klucz już występuje w słowniku, jeżeli tak to nie dodaje nowego.
        print(f'checked row to {checked_row}')
        key = create_key(checked_row)
        print('dodaję key')
        print(f'key to {key}')
        print(f'to jest checked_row przed short key {checked_row}')
        short_key = create_short_key(checked_row)
        print(f'to jest short key  {short_key}')
        part_number = str(sheet.cell(row=row_counter, column=part_col).value)
        part_name = str(sheet.cell(row=row_counter, column=name_col).value)

        print(f'checked_row to: {checked_row}')
        if key not in paths.keys():
            paths[key] = str(paths[short_key]) + '\\' + part_number + ' ' + part_name


    row_counter += 1
    #checked_row = []
    below_row = []

# creating folders
for i, path in paths.items():
    print(f'to jest klucz {i}')
    print(f'to jest ścieżka {path}')
    try:
        if i != '[0]':
            os.makedirs(str(path))
        else:
            print('to jest katalog startowy')
    except FileExistsError:
        print('Ten folder już istnieje')

# copying files
row_counter = 2
checked_row = [0]

print(paths)
print(paths_short)
while row_counter < last_row:
    checked_row = [0]
    for col in range(1, 6):
        if sheet.cell(row=row_counter, column=col).value != None: # sprawdza czy wartość komórki jest None
            checked_row.append(sheet.cell(row=row_counter, column=col).value) # jeżeli nie to dodaje do listy
    drawing = sheet.cell(row_counter, part_col).value
    print(f'to jest {drawing}')
    for file in files_to_copy:
        if drawing in file:
            print("tak jest w liście")
            to_be_copied = str(current_path) + '\\' + str(file)
    destination_folder = paths[create_short_key(checked_row)]
    print(f'to jest to be_copied {to_be_copied}')
    print(f'to jest kopiowany file {drawing}')
    print(str(current_path) + '\\' + str(file))
    shutil.copy(to_be_copied, destination_folder)

    row_counter += 1

#for k, v in paths.items():
#    print(f' {k} \n {v}')

# path length can't be longer than 256 characters
#length_check_list = []
#for v in paths.values():
#    length_check_list.append(v)

#print(len(max(length_check_list, key=len)))
#print(files_to_copy)
